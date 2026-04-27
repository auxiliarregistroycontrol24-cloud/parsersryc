# =============================================================================
# PARCHE: Exportar texto de LLMWhisperer a Excel
# =============================================================================
# Instrucciones de integración al final del archivo.
# =============================================================================

import os
import time
import threading
import queue
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# 1. FUNCIÓN PRINCIPAL: Exportar texto LLMWhisperer a Excel
# ---------------------------------------------------------------------------

def exportar_llmwhisperer_a_excel(
    textos_extraidos: list[dict],
    ruta_salida: str,
) -> str:
    """
    Recibe una lista de dicts con claves:
      - 'archivo': nombre del archivo origen
      - 'texto':   texto crudo devuelto por LLMWhisperer
      - 'paginas': número de páginas detectadas (opcional)
      - 'timestamp': marca de tiempo de extracción (opcional)

    Genera un Excel con:
      - Hoja "Resumen": una fila por archivo con metadatos
      - Hoja "Texto Completo": el texto crudo de cada PDF en su propia sección
    Devuelve la ruta del archivo creado.
    """
    wb = openpyxl.Workbook()

    # --- Estilos reutilizables ---
    COLOR_HEADER    = "1F4E79"   # azul oscuro
    COLOR_SUBHEADER = "2E75B6"   # azul medio
    COLOR_VERDE     = "E2EFDA"   # verde claro (fila par)
    COLOR_BLANCO    = "FFFFFF"

    fuente_titulo   = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    fuente_header   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    fuente_normal   = Font(name="Arial", size=9)
    fuente_mono     = Font(name="Courier New", size=8)

    borde_fino = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def estilo_header(cell, color=COLOR_HEADER):
        cell.font = fuente_header
        cell.fill = PatternFill("solid", start_color=color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borde_fino

    def estilo_celda(cell, color=COLOR_BLANCO, wrap=False):
        cell.font = fuente_normal
        cell.fill = PatternFill("solid", start_color=color)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=wrap)
        cell.border = borde_fino

    # =========================================================================
    # HOJA 1: RESUMEN
    # =========================================================================
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"

    # Título principal
    ws_resumen.merge_cells("A1:F1")
    titulo = ws_resumen["A1"]
    titulo.value = "📄 Reporte de Extracción LLMWhisperer"
    titulo.font = fuente_titulo
    titulo.fill = PatternFill("solid", start_color=COLOR_HEADER)
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws_resumen.row_dimensions[1].height = 28

    # Fecha de generación
    ws_resumen.merge_cells("A2:F2")
    fecha_cell = ws_resumen["A2"]
    fecha_cell.value = f"Generado: {time.strftime('%Y-%m-%d  %H:%M:%S')}   |   Total archivos: {len(textos_extraidos)}"
    fecha_cell.font = Font(name="Arial", italic=True, size=9, color="595959")
    fecha_cell.alignment = Alignment(horizontal="center")
    ws_resumen.row_dimensions[2].height = 16

    # Cabeceras de tabla
    cabeceras = ["#", "Archivo Origen", "Páginas", "Caracteres Extraídos", "Palabras", "Timestamp"]
    anchos    = [5,   45,               10,        22,                      12,         22        ]

    for col_idx, (cab, ancho) in enumerate(zip(cabeceras, anchos), start=1):
        cell = ws_resumen.cell(row=3, column=col_idx, value=cab)
        estilo_header(cell, COLOR_SUBHEADER)
        ws_resumen.column_dimensions[get_column_letter(col_idx)].width = ancho

    ws_resumen.row_dimensions[3].height = 20

    # Filas de datos
    for fila_idx, item in enumerate(textos_extraidos, start=1):
        row_num = fila_idx + 3
        color_fila = COLOR_VERDE if fila_idx % 2 == 0 else COLOR_BLANCO

        texto = item.get("texto", "") or ""
        num_paginas = texto.count("--- NUEVA PÁGINA ---") + 1 if texto else 0
        num_chars   = len(texto)
        num_palabras = len(texto.split()) if texto else 0
        timestamp   = item.get("timestamp", time.strftime("%Y-%m-%d %H:%M:%S"))

        valores = [
            fila_idx,
            item.get("archivo", ""),
            num_paginas,
            num_chars,
            num_palabras,
            timestamp,
        ]
        for col_idx, val in enumerate(valores, start=1):
            cell = ws_resumen.cell(row=row_num, column=col_idx, value=val)
            estilo_celda(cell, color_fila)
            if col_idx in (3, 4, 5):
                cell.alignment = Alignment(horizontal="center", vertical="top")

    # Fila de totales
    total_row = len(textos_extraidos) + 4
    ws_resumen.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=9)
    ws_resumen.cell(row=total_row, column=2, value=f"{len(textos_extraidos)} archivos").font = Font(name="Arial", bold=True, size=9)
    for col in range(1, 7):
        ws_resumen.cell(row=total_row, column=col).fill = PatternFill("solid", start_color="D6E4F0")
        ws_resumen.cell(row=total_row, column=col).border = borde_fino

    ws_resumen.freeze_panes = "A4"

    # =========================================================================
    # HOJA 2: TEXTO COMPLETO
    # =========================================================================
    ws_texto = wb.create_sheet("Texto Completo")

    # Cabecera de hoja
    ws_texto.merge_cells("A1:C1")
    titulo2 = ws_texto["A1"]
    titulo2.value = "Texto Extraído por LLMWhisperer (Texto Crudo)"
    titulo2.font = fuente_titulo
    titulo2.fill = PatternFill("solid", start_color=COLOR_HEADER)
    titulo2.alignment = Alignment(horizontal="center", vertical="center")
    ws_texto.row_dimensions[1].height = 26

    cabeceras2 = ["#", "Archivo Origen", "Texto Extraído"]
    anchos2    = [5,   40,               120             ]
    for col_idx, (cab, ancho) in enumerate(zip(cabeceras2, anchos2), start=1):
        cell = ws_texto.cell(row=2, column=col_idx, value=cab)
        estilo_header(cell, COLOR_SUBHEADER)
        ws_texto.column_dimensions[get_column_letter(col_idx)].width = ancho

    fila_actual = 3
    for num, item in enumerate(textos_extraidos, start=1):
        texto = item.get("texto", "") or "(Sin texto extraído)"

        cell_num = ws_texto.cell(row=fila_actual, column=1, value=num)
        cell_num.font = fuente_normal
        cell_num.alignment = Alignment(horizontal="center", vertical="top")
        cell_num.border = borde_fino
        cell_num.fill = PatternFill("solid", start_color=COLOR_BLANCO)

        cell_archivo = ws_texto.cell(row=fila_actual, column=2, value=item.get("archivo", ""))
        cell_archivo.font = Font(name="Arial", bold=True, size=9)
        cell_archivo.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell_archivo.border = borde_fino
        cell_archivo.fill = PatternFill("solid", start_color="DEEAF1")

        cell_texto = ws_texto.cell(row=fila_actual, column=3, value=texto)
        cell_texto.font = fuente_mono
        cell_texto.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell_texto.border = borde_fino
        cell_texto.fill = PatternFill("solid", start_color=COLOR_BLANCO)

        # Alto de fila proporcional al texto (máx ~400pt para no colapsar Excel)
        lineas = min(texto.count("\n") + 1, 80)
        ws_texto.row_dimensions[fila_actual].height = max(30, lineas * 14)

        fila_actual += 1

    ws_texto.freeze_panes = "A3"

    wb.save(ruta_salida)
    return ruta_salida


# ---------------------------------------------------------------------------
# 2. FUNCIÓN DE LANZAMIENTO EN HILO (para no bloquear la GUI)
# ---------------------------------------------------------------------------

def lanzar_exportacion_llmwhisperer_en_hilo(
    archivos_seleccionados: list[str],
    ventana_principal,
    extraer_texto_pdf_con_apis,       # función del módulo principal
    extraer_texto_excel_con_pandas,   # función del módulo principal
    etiqueta_estado=None,
    barra_progreso=None,
):
    """
    Abre un diálogo para elegir dónde guardar el Excel,
    luego procesa cada archivo con LLMWhisperer en un hilo separado
    y exporta el resultado.
    """
    if not archivos_seleccionados:
        messagebox.showwarning(
            "Sin archivos",
            "No seleccionó ningún archivo para exportar.",
            parent=ventana_principal,
        )
        return

    ruta_salida = filedialog.asksaveasfilename(
        title="Guardar Excel de extracción LLMWhisperer",
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx")],
        initialfile=f"Extraccion_LLMWhisperer_{time.strftime('%Y%m%d_%H%M%S')}.xlsx",
        parent=ventana_principal,
    )
    if not ruta_salida:
        return

    def proceso():
        textos = []
        total = len(archivos_seleccionados)

        for idx, ruta_archivo in enumerate(archivos_seleccionados, start=1):
            nombre_base = os.path.basename(ruta_archivo)
            timestamp_extraccion = time.strftime("%Y-%m-%d %H:%M:%S")

            if etiqueta_estado:
                ventana_principal.after(
                    0,
                    lambda n=nombre_base, i=idx, t=total:
                        etiqueta_estado.config(text=f"Extrayendo con LLMWhisperer: {n} ({i}/{t})"),
                )
            if barra_progreso:
                ventana_principal.after(
                    0,
                    lambda i=idx, t=total:
                        barra_progreso.config(value=(i / t) * 100),
                )

            # Extraer texto según tipo de archivo
            texto_extraido = None
            extension = os.path.splitext(ruta_archivo)[1].lower()
            if extension == ".pdf":
                texto_extraido = extraer_texto_pdf_con_apis(ruta_archivo)
            elif extension in (".xlsx", ".xls"):
                texto_extraido = extraer_texto_excel_con_pandas(ruta_archivo)

            textos.append({
                "archivo":   nombre_base,
                "texto":     texto_extraido or "(No se pudo extraer texto)",
                "timestamp": timestamp_extraccion,
            })

        # Generar Excel
        try:
            exportar_llmwhisperer_a_excel(textos, ruta_salida)
            ventana_principal.after(
                0,
                lambda: messagebox.showinfo(
                    "Exportación Completada",
                    f"✅ Excel generado exitosamente:\n\n{ruta_salida}\n\n"
                    f"Archivos procesados: {len(textos)}",
                    parent=ventana_principal,
                ),
            )
            # Abrir el archivo automáticamente
            ventana_principal.after(0, lambda: _abrir_archivo(ruta_salida))
        except Exception as e:
            ventana_principal.after(
                0,
                lambda: messagebox.showerror(
                    "Error al exportar",
                    f"No se pudo generar el Excel:\n{e}",
                    parent=ventana_principal,
                ),
            )
        finally:
            if etiqueta_estado:
                ventana_principal.after(
                    0, lambda: etiqueta_estado.config(text="Exportación finalizada.")
                )
            if barra_progreso:
                ventana_principal.after(0, lambda: barra_progreso.config(value=0))

    threading.Thread(target=proceso, daemon=True).start()


def _abrir_archivo(ruta):
    import subprocess, os
    try:
        if os.name == "nt":
            os.startfile(ruta)
        elif os.name == "posix":
            subprocess.run(["xdg-open", ruta])
    except Exception:
        pass


# =============================================================================
# INSTRUCCIONES DE INTEGRACIÓN
# =============================================================================
"""
PASO 1 — Copiar este archivo
    Coloca este archivo (parche_exportar_llmwhisperer.py) en la misma carpeta
    que tu script principal.

PASO 2 — Importar al inicio del script principal
    Agrega esta línea junto a los otros imports:

        from parche_exportar_llmwhisperer import lanzar_exportacion_llmwhisperer_en_hilo

PASO 3 — Instalar dependencia (si no tienes openpyxl ya instalado)
    pip install openpyxl

PASO 4 — Agregar el botón en la GUI
    Dentro de la función crear_ventana_principal() (o donde construyes los botones),
    agrega el botón nuevo justo después de boton_seleccionar_gui:

        boton_exportar_llm_gui = ttk.Button(
            botones_frame,
            text="📊 Exportar LLMWhisperer → Excel",
            command=lambda: _seleccionar_y_exportar_llmwhisperer(
                ventana_principal_app,
                etiqueta_estado_gui,
                barra_progreso_gui,
            ),
        )
        boton_exportar_llm_gui.pack(pady=5)

PASO 5 — Agregar la función helper _seleccionar_y_exportar_llmwhisperer
    También dentro de crear_ventana_principal(), agrega esta función auxiliar
    (antes del mainloop):

        def _seleccionar_y_exportar_llmwhisperer(ventana, etiqueta, barra):
            archivos = filedialog.askopenfilenames(
                title="Seleccione PDFs o Excel para extraer con LLMWhisperer",
                filetypes=[
                    ("Documentos soportados", "*.pdf *.xlsx *.xls"),
                    ("PDF", "*.pdf"),
                    ("Excel", "*.xlsx *.xls"),
                ],
                parent=ventana,
            )
            if archivos:
                lanzar_exportacion_llmwhisperer_en_hilo(
                    list(archivos),
                    ventana,
                    extraer_texto_pdf_con_apis,
                    extraer_texto_excel_con_pandas,
                    etiqueta_estado=etiqueta,
                    barra_progreso=barra,
                )

FIN — Con estos 5 pasos el botón aparecerá en la interfaz y generará el Excel
      con el texto crudo de LLMWhisperer sin pasar por ningún modelo de IA.
"""
